from cmatrix import *
import numpy as np
import pandas as pd
import pyexcel as p
import openpyxl as o
import array_to_latex as a
import re
import os

# sudo apt install tesseract-ocr libtesseract-dev tesseract-ocr-eng

def Matrix2List(Mat):
	"""
		Converting a Matrix objedct to its associated Data list. The size is appended at the end.

		=============== ========= ========================
		**Parameters**   **Type**   **Description**
		*Mat*             Matrix    The Matrix to convert
		=============== ========= ========================

		:returns: List : The converted matrix as data list
	"""
	data=[]
	for i in range(0,Mat.size_r()):
		for j in range(0,Mat.size_c()):
			data.append(Mat[i,j])
	data.append(Mat.size_r())
	data.append(Mat.size_c())
	return data

def List2Matrix(List,size_r,size_c):
	"""
		Converting a list and Known sizes to a Matrix object.

		================ ========== =======================
		**Parameters**    **Type**   **Description**
		*List*            List       The List to convert
		*size_r*          Int        The number of Rows
		*size_c*          Int        The number of Columns
		================ ========== =======================

		:returns: Matrix : A Matrix Object from size (size_r x size_c) containing the data list.
	"""
	data=Matrix(size_r,size_c)
	ind=0
	for i in range(0,Mat.size_r()):
		for j in range(0,Mat.size_c()):
			data[i,j]=List[ind]
			ind+=1
	return data

def Matrix2Numpy(Mat):
	"""
		Converting a Matrix object to a Numpy Array.

		============== ========== =======================
		**Parameters**  **Type**   **Description**
		*Mat*            Matrix    The Matrix to convert
		============== ========== =======================

		:returns: Numpy Array: The Numpy Array representing the Matrix Object
	"""
	data=[]
	for i in range(0,Mat.size_r()):
		data.append(Mat.line(i))
	return np.array(data)

def Numpy2Matrix(Array):
	"""
		Converting a Numpy Array to a Matrx Object.

		============== ============== ===========================
		**Parameters**   **Type**     **Description**
		*Array*          Numpy Array  The Numpy Array to convert
		============== ============== ===========================

		:returns: Matrix: The Matrix Object representing the Numpy Array.
	"""
	data=Matrix(len(Array),len(Array[0]))
	for i in range(0,Mat.size_r()):
		for j in range(0,Mat.size_c()):
			data[i,j]=Array[i][j]
	return data

def Matrix2Panda(Mat):
	"""
		Converting a Matrix Object to a Pandas DataFrame.

		=============== =========== =======================
		**Parameters**    **Type**   **Description**
		*Mat*              Matrix    The Matrix to convert
		=============== =========== =======================

		:returns: Pandas DataFrame : The DataFrame representing the Matrix Object

	"""
	nparray=Matrix2Numpy(Mat)
	return pd.DataFrame(nparray,dtype='float')

def Panda2Matrix(df):
	"""
		Converting a Pandas DataFrame to a Matrix Object.

		=============== =================== ==========================
		**Parameters**    **Type**          **Description**
		*df*              Pandas DataFrame   The DataFrame to convert
		=============== =================== ==========================

		:returns: Matrix : The Matrix Object representing the DataFrame
	"""
	return Numpy2Matrix(df.to_numpy(dtype='float'))

def Matrix2Ods(Mat,name):
	"""
		Converting a Matrix Object to an Table ods file.

		=============== ========== =============================
		**Parameters**    **Type**   **Description**
		*Mat*             Matrix     The Matrix to convert
		*name*            String     The name of the .ods file
		=============== ========== =============================

		:returns: None
	"""
	to_write=[]
	for i in range(0,Mat.size_r()):
		to_write.append(Mat.line(i))
	Content={'Sheet 1':to_write}
	book=p.Book(Content)
	book.save_as(name+'.xlsx')
	os.system('soffice --headless --convert-to ods *.xlsx')
	os.system('rm *.xlsx')

def Ods2Matrix(pathname):
	"""
		Converting a Table ods file to a Matrix Object.

		================ ======== ===============================================================
		**Parameters**   **Type**   **Description**
		*pathname*       String     The exact name of the file to read, path should be indicated
		================ ======== ===============================================================

		:returns: Matrix: The Matrix containing all the .ods data
	"""
	Records=p.get_array(file_name=pathname)
	size_r=len(Records[0])
	size_c=0
	for i in range(0,len(Records)):
		if(Records[i][0]!=''):
			size_c+=1
	Mat=Matrix(size_r,size_c)
	for i in range(0,Mat.size_r()):
		for j in range(0,Mat.size_c()):
			Mat[i,j]=Records[i][j]
	return Mat

def Matrix2xlsx(Mat,name):
	"""
		Converting a Matrix Object to a Excel File.

		=============== ========== ============================
		**Parameters**   **Type**   **Description**
		*Mat*            Matrix     The Matrix to convert
		*name*           String     The name of the .xlsx file
		=============== ========== ============================

		:returns: None
	"""
	Matrix2Ods(Mat,name)
	os.system('soffice --headless --convert-to xlsx '+name+'.ods')
	os.system('rm '+name+'.ods')

def xlsx2Matrix(pathname):
	"""
		Converting an Excel File to a Matrix Object.

		=============== ========== ==============================================================
		**Parameters**   **Type**   **Description**
		*pathname*       String     The exact name of the file to read, path should be indicated
		=============== ========== ==============================================================

		:returns: Matrix: The Matrix containing all the .xlsx data
	"""
	Wb=o.load_workbook(filename=pathname)
	Ws=Wb.worksheets[0]
	data=Matrix(Ws.max_row,Ws.max_column)
	i,j = 0,0
	for line in Ws.rows:
		for item in line:
			data[i,j]=item.value
			i+=1
		i=0
		j+=1
	return data.transpose()

def Matrix2tex(Mat,name):
	"""
		Converting a Matrix Object to a Tex Matrix.

		============== =========== ===========================
		**Parameters**   **Type**   **Description**
		*Mat*            Matrix     The Matrix to convert
		*name*           String     The name of the .tex file
		============== =========== ===========================

		:returns:  None
	"""
	npa=Matrix2Numpy(Mat)
	latex_code='\\documentclass{article}\n\\usepackage{nicematrix}\n\\begin{document}\n$'
	latex_code+=a.to_ltx(npa, frmt = '{:6.2f}', arraytype = 'bmatrix', print_out=False)
	latex_code=latex_code+'$\n\\end{document}'
	f=open(name+'.tex','w')
	f.write(latex_code)

def tex2Matrix(pathname):
	"""
		Converting a Tex Matrix to a Matrix Object (Only Matrix object will be treated).

		=============== ========== =============================================================
		**Parameters**   **Type**   **Description**
		*pathname*       String     The exact name of the file to read, path should be indicated
		=============== ========== =============================================================

		:returns: Matrix: The Matrix containing the .tex data
	"""
	f=open(pathname,'r')
	Content=f.readlines()
	regex=r'[0-9]+\.[0-9]*'
	data=[]
	size_c=0
	size_r=0
	for line in Content:
		new_line=[]
		matches=re.finditer(regex, line, re.MULTILINE)
		for matchNum, match in enumerate(matches, start=1):
			new_line.append(float(match.group()))
		if (new_line != []):
			data.append(new_line)
		size_c+=1
	size_r=len(data[0])
	size_c-=2
	return Numpy2Matrix(data)

def Matrix2png(Mat,name):
	"""
		Converting a Matrix Object to a png picture.

		============== =========== ===========================
		**Parameters**   **Type**   **Description**
		*Mat*            Matrix     The Matrix to convert
		*name*           String     The name of the .png file
		============== =========== ===========================
	"""
	Matrix2tex(Mat,name)
	os.system('pdflatex '+name+'.tex')
	os.system('pdftoppm '+name+'.pdf '+name+'.pdf -png')
	os.system('rm '+name+'.pdf')

def png2Matrix(name):
	"""
		Converting a Png picture to a Matrix Object (Only Matrix object will be treated).

		=============== ========== ====================================
		**Parameters**   **Type**   **Description**
		*name*           String     The exact name of the file to read
		=============== ========== ====================================

		:returns: Matrix: The Matrix containing all the Digital Informations of the png file.
	"""
	os.system('tesseract -l eng '+name+' '+name)
	Contents=os.popen('cat *.txt').readlines()
	data=[]
	Wline=[]

	for line in Contents:
		try:
			# print(line)
			line.replace('\n','')
			tmp=(line.split(' '))
			for item in tmp:
				Wline.append(float(item))
			data.append(Wline)
			Wline=[]
			print(data)
		except:
			pass
	return Numpy2Matrix(np.array(data))

def Matrix2pdf(Mat,name):
	"""
		Converting a Matrix Object to a PDF File.

		=============== ========== ============================
		**Parameters**   **Type**   **Description**
		*Mat*            Matrix     The Matrix to convert
		*name*           String     The name of the .pdf file
		=============== ========== ============================

		:returns: None 
	"""
	Matrix2Ods(Mat,name)
	os.system('soffice --headless --convert-to pdf:calc_pdf_Export '+name+'.ods')
	os.system('rm '+name+'.ods')

# Testing Procedure
Mat=rand(4)
print(Mat)
print(Matrix2List(Mat))
print(List2Matrix(Matrix2List(Mat),Mat.size_r(),Mat.size_c()))
print(Matrix2Numpy(Mat))
print(Numpy2Matrix(Matrix2Numpy(Mat)))
print(Matrix2Panda(Mat))
print(Panda2Matrix(Matrix2Panda(Mat)))
Matrix2Ods(Mat,'test')
print(Ods2Matrix('test.ods'))
Matrix2xlsx(Mat,'test')
print(xlsx2Matrix('test.xlsx'))
Matrix2tex(Mat,'test')
print(tex2Matrix('test.tex'))
Matrix2png(Mat,'test')
print(png2Matrix('test.pdf-1.png'))
Matrix2pdf(Mat,'test')
os.system('rm test*')